import os
import smtplib
import mimetypes
from string import Template
from email.message import EmailMessage

from PySide6.QtCore import QThread, Signal
from csv import DictReader, __version__ as csv_version
from openpyxl import load_workbook, __version__ as openpyxl_version
import html2text

from utils.logger import logger


class EmailerThread(QThread):
    outputSignal = Signal(str, str)

    def __init__(
        self,
        smtpHost: str,
        smtpPort: str,
        smtpUsername: str,
        smtpPassword: str,
        smtpDelay: int,
        subject: str,
        origin: str,
        reply: str,
        cc: str,
        bcc: str,
        bodyPath: str,
        tablePath: str,
        attachmentPath: str,
    ):
        super().__init__()
        self.smtpHost = smtpHost
        self.smtpPort = smtpPort
        self.smtpUsername = smtpUsername
        self.smtpPassword = smtpPassword
        self.smtpDelay = smtpDelay
        self.subject = subject
        self.origin = origin
        self.reply = reply
        self.cc = cc
        self.bcc = bcc
        self.bodyPath = bodyPath
        self.tablePath = tablePath
        self.attachmentPath = attachmentPath

    def output(self, text: str, level: str = "INFO"):
        logger.log(level, text)
        self.outputSignal.emit(text, level)

    def readBody(self) -> Template:
        with open(self.bodyPath, "r", encoding="utf-8") as bodyFile:
            content = bodyFile.read()
        return Template(content)

    def readTable(self) -> tuple[list[dict[str, str]], list[str]]:
        if self.tablePath.endswith(".csv"):
            logger.info(
                f"Using native CSV {csv_version} module to read {self.tablePath}"
            )
            records = []
            with open(self.tablePath, mode="r", encoding="utf-8") as csvfile:
                reader = DictReader(csvfile)
                headers = reader.fieldnames
                if not headers:
                    raise ValueError("CSV file has no headers")
                logger.info(f"Headers found in CSV: {len(headers)}")
                for row in reader:
                    records.append(
                        {
                            key: (value if value else "")
                            for key, value in row.items()
                        }
                    )
        elif self.tablePath.endswith(".xlsx"):
            logger.info(
                f"Using openpyxl {openpyxl_version} to read {self.tablePath}"
            )
            workbook = load_workbook(filename=self.tablePath, data_only=True)
            logger.info(
                f"Loaded workbook with {len(workbook.sheetnames)} sheets"
            )
            sheet = workbook.active
            if not sheet:
                raise ValueError("Excel file has no sheets")
            logger.info(
                f"Active sheet: {sheet.title}, {sheet.max_row} rows, {sheet.max_column} columns"
            )
            headers = [str(cell.value) for cell in sheet[1] if cell.value]
            logger.info(f"Headers found: {len(headers)}")
            if not headers:
                raise ValueError("Excel file has no headers")
            records = []
            for row in sheet.iter_rows(min_row=2, values_only=True):  # type: ignore
                logger.debug(f"Row data: {row}")
                record = {
                    headers[i]: (value if value else "")
                    for i, value in enumerate(row)
                    if i < len(headers)
                }
                records.append(record)
        else:
            raise ValueError("Unsupported file extension")

        return records, list(headers)

    def readAttachment(self, filePath: str) -> tuple[bytes, str, str]:
        mimeType, _ = mimetypes.guess_type(filePath)

        if mimeType is None:
            mimeType = "application/octet-stream"
        mainType, subType = mimeType.split("/")

        with open(filePath, "rb") as file:
            content = file.read()

        return content, mainType, subType

    def run(self):
        inputs = self.__dict__.copy()
        inputs.pop("smtpUsername", False)
        inputs.pop("smtpPassword", False)

        logger.info(f"Starting emailer thread with input parameters: {inputs}")

        self.output("...")
        with logger.catch():
            if not os.path.isdir(self.attachmentPath):
                self.output(
                    f"Attachment directory {self.attachmentPath} not found or is not a directory",
                    "ERROR",
                )
                return

            try:
                body = self.readBody()
            except Exception as e:
                self.output(f"Failed to load body: {e}", "ERROR")
                return
            logger.info(f"Body loaded: {body.template}")
            logger.info(f"Body read: {body.template}")

            try:
                rows, cols = self.readTable()
            except Exception as e:
                self.output(f"Failed to load table: {e}", "ERROR")
                return
            logger.info(
                f"Table loaded, found {len(rows)} {len(rows) == 1 and 'row' or 'rows'}"
            )
            logger.info(f"Table columns read: {cols}")
            logger.info(f"Table rows read: {rows}")

            if not rows or not cols:
                self.output("Given table is empty", "ERROR")
                return
            if len(rows) < 1:
                self.output(
                    "Given table must have at least another row in addition to the headers!",
                    "ERROR",
                )
                return

            variableInputs = {
                "Subject": self.subject,
                "CC": self.cc,
                "BCC": self.bcc,
                "Body": self.attachmentPath,
            }
            # Validate that all variable inputs are present in the table headers
            for key, value in variableInputs.items():
                try:
                    Template(value).substitute(rows[0])
                except KeyError as e:
                    self.output(
                        f"A variable interpolation key used in '{key}' was not found in the table headers: {e}",
                        "ERROR",
                    )
                    return
                except TypeError as e:
                    self.output(
                        f"A variable interpolation key used in '{key}' was found more than once in table headers: {e}",
                        "ERROR",
                    )
                    return
                except ValueError as e:
                    self.output(
                        f"A variable interpolation key used in '{key}' is empty or invalid, make sure you avoid special characters in the variable name: {e}",
                        "ERROR",
                    )
                    return

            server = smtplib.SMTP(host=self.smtpHost, port=int(self.smtpPort))

            try:
                server.starttls()
                server.login(self.smtpUsername, self.smtpPassword)
                self.output(
                    f"Logged in as {self.smtpUsername} on {self.smtpHost}:{self.smtpPort}"
                )
            except Exception as e:
                self.output(f"Failed to login: {e}", "ERROR")
                return

            total = 0
            successful = 0
            for row in rows:
                total += 1
                row_number = total + 1  # Exclude header row

                message = EmailMessage()

                if not row[cols[0]] or "@" not in row[cols[0]]:
                    self.output(
                        f"[{row_number}] Email address column is empty or the value is invalid on this row, you should either fix it or delete it",
                        "ERROR",
                    )
                    continue

                # Configure email headers
                message["To"] = row[cols[0]].strip()
                message["From"] = self.origin.strip()
                message["Subject"] = (
                    Template(self.subject).substitute(row).strip()
                )
                if self.reply:
                    message["Reply-To"] = (
                        Template(self.reply).substitute(row).strip()
                    )
                if self.cc:
                    message["Cc"] = Template(self.cc).substitute(row).strip()
                if self.bcc:
                    message["Bcc"] = Template(self.bcc).substitute(row).strip()

                # Add body to email
                content = body.substitute(row)
                if self.bodyPath.endswith(".html"):
                    message.set_content(html2text.html2text(content))
                    message.add_alternative(content, subtype="html")
                else:
                    message.set_content(content)

                # Add attachments to email
                filePaths: list[str] = []
                try:
                    if len(cols) > 1 and row[cols[1]]:
                        attachmentNames = row[cols[1]].split("#")
                        for attachmentName in attachmentNames:
                            fullPath = os.path.join(
                                self.attachmentPath, attachmentName
                            )
                            filePaths.append(fullPath)
                    for filePath in filePaths:
                        file, mainType, subType = self.readAttachment(filePath)
                        message.add_attachment(
                            file,
                            maintype=mainType,
                            subtype=subType,
                            filename=os.path.basename(filePath),
                        )
                except Exception as e:
                    self.output(
                        f"[{row_number}] Failed to attach file in email to {row[cols[0]]}: {e}",
                        "ERROR",
                    )
                    continue

                # Send email
                try:
                    server.send_message(message)
                    successful += 1
                    self.output(
                        f"[{row_number}] Sent email to {row[cols[0]]} with {len(filePaths)} {len(filePaths) == 1 and 'attachment' or 'attachments'}"
                    )
                    self.output(
                        f"Waiting {self.smtpDelay} seconds before sending the next email to avoid server rate limits..."
                    )
                    self.msleep(self.smtpDelay * 1000)

                except Exception as e:
                    self.output(
                        f"[{row_number}] Failed to send email to {row[cols[0]]}: {e}",
                        "ERROR",
                    )

            server.quit()

            self.output(f"Successfully sent {successful} out of {total} emails")
